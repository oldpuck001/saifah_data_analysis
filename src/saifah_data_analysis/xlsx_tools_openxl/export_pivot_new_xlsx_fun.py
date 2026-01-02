# export_pivot_new_xlsx_fun.py

# 导出新数据透视表 xlsx 文件

from tkinter import filedialog
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def export_pivot_new_xlsx(df):

    # 实例化工作簿
    wb = openpyxl.Workbook()

    # 获取第一张工作表并赋予它一个名称
    ws = wb.active

    # 重設索引並將DataFrame的數據添加到Excel工作表
    df.reset_index(inplace=True)

    # 修改DataFrame的第1列第1行（列标题）为空白
    df.columns.values[0] = ''

    # 将DataFrame的数据添加到Excel工作表
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # 添加合计公式到最后一列
    colnumber = ws.max_column + 1
    row_sum_title = ws.cell(row=1, column=colnumber)
    row_sum_title.value = 'Total'
    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
        row_sum_cell = ws.cell(row=row_index + 2, column=colnumber)
        row_sum_cell.value = f'=SUM({row[1].coordinate}:{row[-2].coordinate})'

    # 添加合计公式到最后一行
    rows_count = ws.max_row + 1
    col_sum_title = ws.cell(row=rows_count, column=1)
    col_sum_title.value = 'Total'
    for col_index in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col_index)
        col_sum_cell = ws.cell(row=rows_count, column=col_index)
        col_sum_cell.value = f'=SUM({col_letter}2:{col_letter}{ws.max_row-1})'

    # 调整列宽
    for column_cells in ws.columns:
        length = max(len(str(cell.value))+8 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length
    
    # 保存工作簿会在磁盘上创建文件
    file_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')])
    wb.save(file_path)

    return f'The xlsx output file path: {file_path}\nThe xlsx output file was exported successfully!\n'