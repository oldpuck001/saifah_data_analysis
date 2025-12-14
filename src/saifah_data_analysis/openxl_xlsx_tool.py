# openxl_xlsx_tool.py

import openpyxl
import openpyxl.styles
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class openxl_xlsx_tool_class:

    # 新建工作簿
    def create_new_xlsx(self, file_path, sheet_name_list):
        # 实例化工作簿
        wb = openpyxl.Workbook()

        # 获取第一张工作表并赋予它一个名称
        ws_1 = wb.active
        ws_1.title = sheet_name_list[0]

        # 获取第二张工作表并赋予它一个名称
        for t in sheet_name_list[1:]:
            wb.create_sheet(title=t)

        # 保存工作簿会在磁盘上创建文件
        wb.save(file_path)

        return f'The xlsx output file path: {file_path}\nThe xlsx output file was created successfully!\n'


    # 逐个填写工作表单元格
    def fill_in_xlsx_cell(self, file_path, fill_in_xlsx):

        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb[fill_in_xlsx[0][0]]

        # 起始单元格位置
        r = fill_in_xlsx[0][1]
        c = fill_in_xlsx[0][2]

        # 设置行高
        r_n = r + 1
        for n in fill_in_xlsx[1]:
            ws.row_dimensions[r_n].height = n
            r_n += 1

        # 设置列宽
        c_n = c
        for n in fill_in_xlsx[2]:
            ws.column_dimensions[get_column_letter(c_n)].width = n
            c_n += 1

        # 填写单元格、设置单元格格式
        r_f = r
        for sheet_row in fill_in_xlsx[3:]:
            r_f += 1
            c_f = c
            for sheet_cell in sheet_row:
                cell_fill = ws.cell(row=r_f, column=c_f, value=sheet_cell[0])
                self.cell_format(cell_fill, sheet_cell[1])
                c_f += 1

        wb.save(file_path)

        return 'The xlsx output file was updata successfully!\n'


    # 填写整个工作表
    def fill_in_xlsx_sheet(self, file_path, fill_sheet, fill_df):

        wb = load_workbook(file_path)
        ws = wb[fill_sheet]

        # 写入数据
        header = fill_df.columns.tolist()

        ws.append(header)
        for row in fill_df.values:
            ws.append(row.tolist())

        wb.save(file_path)

        return 'The xlsx output file was updata successfully!\n'


    # 从指定位置开始填写整个工作表
    def fill_in_xlsx_sheet_cell(self, file_path, fill_sheet, fill_df, start_row, start_col):

        wb = load_workbook(file_path)
        ws = wb[fill_sheet]

        for j, col in enumerate(fill_df.columns):
            ws.cell(row=start_row, column=start_col + j, value=col)
        data_start = start_row + 1

        for i, row in enumerate(fill_df.values):
            for j, value in enumerate(row):
                ws.cell(row=data_start+i, column=start_col+j, value=value)

        wb.save(file_path)

        return 'The xlsx output file was updata successfully!\n'


    # 单元格格式
    def cell_format(self, cell_fill, n):

        # 边框格式
        thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                            right=openpyxl.styles.Side(style='thin'),
                                            top=openpyxl.styles.Side(style='thin'),
                                            bottom=openpyxl.styles.Side(style='thin'))

        if n == 1:
            cell_fill.border = thin_border
            cell_fill.alignment = Alignment(horizontal='center', vertical='center')

        elif n == 2:
            cell_fill.border = thin_border
            cell_fill.alignment = Alignment(vertical='center')

        elif n == 3:
            cell_fill.border = thin_border
            cell_fill.alignment = Alignment(horizontal='center', vertical='center')

        elif n == 4:
            cell_fill.border = thin_border
            cell_fill.alignment = Alignment(vertical='center')
            cell_fill.number_format = '#,##0.00'

        return